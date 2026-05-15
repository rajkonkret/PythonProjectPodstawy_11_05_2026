import openpyxl

# pip install openpyxl

workbook = openpyxl.load_workbook('dane.xlsx')
print(workbook)  # <openpyxl.workbook.workbook.Workbook object at 0x00000215694EF4D0>

worksheet = workbook.active
print(worksheet)  # <Worksheet "Sheet1">

for i in worksheet:
    print(i)
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>, <Cell 'Sheet1'.D1>)
# (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>, <Cell 'Sheet1'.D2>)
# (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>, <Cell 'Sheet1'.D3>)

lista = []
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)

print(lista)
# [None, 'Sales Date', 'Sales Person', 'Amount'
# 0, datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000,

# slicowanie
for i in range(0, len(lista), 4):  # (start, stop, krok) u nas krok=4
    print(lista[i:i + 4])  # [start:stop] [0:4] 0123 None, 'Sales Date', 'Sales Person', 'Amount'

# [None, 'Sales Date', 'Sales Person', 'Amount']
# [0, datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000]
# [1, datetime.datetime(2019, 12, 6, 0, 0), 'Mir Hossain', 50000]
